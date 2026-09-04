# -*- coding: utf-8 -*-
"""
Гелиус Синерджи — синхронизация таблицы агентов со страницей.

  Гелиус_агенты.xlsx  ──sync.py──▶  index.html  ──sync.cmd──▶  GitHub Pages

Читает лист «Агенты», проверяет его и перезаписывает блок между маркерами
«АГЕНТЫ · НАЧАЛО» и «АГЕНТЫ · КОНЕЦ» в index.html. Добавили строку в таблице —
на карте появился агент; удалили строку — агент исчез.

Ничего, кроме этого блока, скрипт не трогает. Категории и стадии он читает
из самой страницы, поэтому переименование категории не ломает синхронизацию.

Запуск:  python sync.py            — перенести правки
         python sync.py --check    — только проверить, ничего не писать
"""
import io
import json
import os
import re
import subprocess
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
PAGE = os.path.join(HERE, 'index.html')
BOOK = os.path.join(HERE, 'Гелиус_агенты.xlsx')

BEGIN = '/* ==== АГЕНТЫ · НАЧАЛО'
END = '/* ==== АГЕНТЫ · КОНЕЦ ==== */'

# Столбцы листа «Агенты» → поля агента на странице.
COLS = ['id', 'cat', 'stage', 'n', 'role', 'pr', 'ac', 'ef', 'ev', 'attn']
FIELDS = ['id', 'c', 's', 'n', 'role', 'pr', 'ac', 'ef', 'ev', 'attn']

ID_RE = re.compile(r'^[a-z0-9][a-z0-9_-]{1,31}$')


def fail(msg):
    print('ОШИБКА: ' + msg)
    sys.exit(1)


def read_page_refs():
    """Категории и стадии берём из страницы — она источник истины для ключей."""
    src = io.open(PAGE, encoding='utf-8').read()
    start = src.find("var NB = '\\u00A0';")
    stop = src.find(BEGIN)
    if start < 0 or stop < 0:
        fail('в index.html не найден блок данных или маркер «АГЕНТЫ · НАЧАЛО»')
    js = src[start:stop] + '\n return {CATS: CATS, STAGES: STAGES};'
    node = subprocess.run(
        ['node', '-e',
         'let f=new Function(process.argv[1]);'
         'process.stdout.write(JSON.stringify(f()));', js],
        capture_output=True)
    if node.returncode != 0:
        fail('не удалось прочитать справочники из index.html:\n'
             + node.stderr.decode('utf-8', 'replace'))
    d = json.loads(node.stdout.decode('utf-8'))
    cats = {c['n'].strip().upper(): c['k'] for c in d['CATS']}
    stages = {s['n'].strip().upper(): s['k'] for s in d['STAGES']}
    return src, cats, stages


def read_book(cats, stages):
    try:
        from openpyxl import load_workbook
    except ImportError:
        fail('нужен openpyxl:  pip install openpyxl')
    if not os.path.exists(BOOK):
        fail('не найдена таблица ' + BOOK)
    wb = load_workbook(BOOK, data_only=True)
    if 'Агенты' not in wb.sheetnames:
        fail('в таблице нет листа «Агенты»')
    ws = wb['Агенты']

    head = [(c.value or '').strip() if isinstance(c.value, str) else c.value
            for c in ws[1]]
    if not head or str(head[0]).strip().lower() != 'id':
        fail('первый столбец листа «Агенты» должен называться id')

    rows, seen, problems = [], {}, []
    for n in range(2, ws.max_row + 1):
        vals = []
        for i in range(len(COLS)):
            v = ws.cell(row=n, column=i + 1).value
            vals.append('' if v is None else str(v).strip())
        if not any(vals):
            continue                      # пустая строка — просто пропускаем

        rec = dict(zip(COLS, vals))
        aid = rec['id']

        if not aid:
            problems.append('строка %d: пустой id' % n)
            continue
        if not ID_RE.match(aid):
            problems.append('строка %d: id «%s» — только латиница, цифры, дефис '
                            'и подчёркивание, 2–32 знака' % (n, aid))
            continue
        if aid in seen:
            problems.append('строка %d: id «%s» уже был в строке %d'
                            % (n, aid, seen[aid]))
            continue
        seen[aid] = n

        ck = cats.get(rec['cat'].upper())
        if not ck:
            problems.append('строка %d: категория «%s» не из справочника (%s)'
                            % (n, rec['cat'], ', '.join(sorted(cats))))
            continue
        sk = stages.get(rec['stage'].upper())
        if not sk:
            problems.append('строка %d: стадия «%s» не из справочника (%s)'
                            % (n, rec['stage'], ', '.join(sorted(stages))))
            continue
        if not rec['n']:
            problems.append('строка %d: пустое название агента' % n)
            continue

        out = {'id': aid, 'c': ck, 's': sk, 'n': rec['n']}
        for f in ('role', 'pr', 'ac', 'ef', 'ev', 'attn'):
            if rec[f]:
                out[f] = rec[f]
        rows.append((n, out))

    if problems:
        print('Таблица не принята, правки не переносились:')
        for p in problems:
            print('  · ' + p)
        sys.exit(1)
    if not rows:
        fail('на листе «Агенты» нет ни одной строки')
    return rows


def main():
    check = '--check' in sys.argv
    src, cats, stages = read_page_refs()
    rows = read_book(cats, stages)

    was = len(re.findall(r'\{"id":', src[src.find(BEGIN):src.find(END)]))
    body = ',\n'.join(json.dumps(r, ensure_ascii=False, sort_keys=False)
                      for _, r in rows)
    block = (
        BEGIN + ' ======================================================\n'
        '   Блок генерируется из Гелиус_агенты.xlsx командой sync.cmd.\n'
        '   Правьте таблицу, а не этот массив: ручные правки затрутся '
        'синхронизацией.\n'
        '   Добавили строку в таблице — на карте появился агент; '
        'удалили — исчез. */\n'
        'var A = [\n' + body + '\n];\n' + END
    )

    i = src.find(BEGIN)
    j = src.find(END) + len(END)
    new = src[:i] + block + src[j:]

    thin = sum(1 for _, r in rows
               if not (r.get('pr') and r.get('ac') and r.get('ef')))
    by_cat, by_stage = {}, {}
    for _, r in rows:
        by_cat[r['c']] = by_cat.get(r['c'], 0) + 1
        by_stage[r['s']] = by_stage.get(r['s'], 0) + 1

    print('Таблица принята.')
    print('  агентов было: %d, стало: %d' % (was, len(rows)))
    print('  по линейкам:  ' + ', '.join('%s %d' % kv for kv in by_cat.items()))
    print('  по стадиям:   ' + ', '.join('%s %d' % kv for kv in by_stage.items()))
    print('  без полного описания (пунктирный обод): %d' % thin)

    if check:
        print('\n--check: страница не изменена.')
        return

    if new == src:
        print('\nСтраница уже соответствует таблице, файл не перезаписан.')
    else:
        io.open(PAGE, 'w', encoding='utf-8').write(new)
        print('\nindex.html обновлён.')

    if '--deploy' in sys.argv:
        deploy()


def deploy():
    """Копирует страницу в репозиторий Pages и публикует."""
    import shutil
    repo = os.path.abspath(os.path.join(HERE, '..', 'helius-pages'))
    if not os.path.isdir(os.path.join(repo, '.git')):
        fail('не найден репозиторий Pages: ' + repo)

    shutil.copyfile(PAGE, os.path.join(repo, 'index.html'))
    print('\nСкопировано в ' + repo)

    def git(*args, **kw):
        return subprocess.run(('git',) + args, cwd=repo,
                              capture_output=True, **kw)

    git('add', '-A')
    if git('diff', '--cached', '--quiet').returncode == 0:
        print('Публиковать нечего: изменений нет.')
        return

    r = git('commit', '-q', '-m',
            'Обновление агентов из Гелиус_агенты.xlsx')
    if r.returncode != 0:
        fail('коммит не прошёл:\n' + r.stderr.decode('utf-8', 'replace'))
    r = git('push', '-q')
    if r.returncode != 0:
        fail('пуш не прошёл:\n' + r.stderr.decode('utf-8', 'replace'))

    print('Опубликовано: https://stepansky0012.github.io/helius-synergy/')
    print('Сборка Pages занимает до двух минут.')


if __name__ == '__main__':
    main()
